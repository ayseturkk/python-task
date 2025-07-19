import argparse
import requests
import sys
import os
import csv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import Optional

SERVER_URL = "http://localhost:8000/vehicles"

def parse_args():
    parser = argparse.ArgumentParser(description="Client for vehicle data processing.")
    parser.add_argument('-f', '--file', type=str, default='vehicles.csv', help='Path to the CSV file to upload (default: vehicles.csv)')
    parser.add_argument('-k', '--keys', nargs='*', default=[], help='Extra columns to include')
    parser.add_argument('-c', '--colored', action=argparse.BooleanOptionalAction, default=True, help='Enable row coloring (default: True)')
    return parser.parse_args()

def send_csv_to_server(csv_path: str) -> list:
    with open(csv_path, 'rb') as f:
        files = {'file': (os.path.basename(csv_path), f, 'text/csv')}
        response = requests.post(SERVER_URL, files=files)
    if response.status_code != 200:
        print(f"Server error: {response.status_code} {response.text}")
        sys.exit(1)
    return response.json()

def sort_vehicles(vehicles: list) -> list:
    return sorted(vehicles, key=lambda v: v.get('gruppe', ''))

def get_hu_color(hu_str: str) -> Optional[str]:
    try:
        hu_date = datetime.strptime(hu_str, "%Y-%m-%d")
    except Exception:
        return None
    now = datetime.now()
    delta = now - hu_date
    if delta.days <= 90:
        return "007500"  # green
    elif delta.days <= 365:
        return "FFA500"  # orange
    else:
        return "b30000"  # red

def write_excel(vehicles: list, keys: list, colored: bool, csv_file: str):
    wb = Workbook()
    ws = wb.active  # type: ignore
    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    ws.title = base_name  # type: ignore

    # Determine columns
    if keys:
        columns = keys
    elif vehicles:
        columns = list(vehicles[0].keys())
    else:
        columns = []
    ws.append(columns)  # type: ignore

    for row_idx, v in enumerate(vehicles, start=2):
        row = [v.get(col, '') for col in columns]
        ws.append(row)  # type: ignore
        # Tint labelIds cell if present
        if 'labelIds' in columns and 'labelColors' in v and v.get('labelColors'):
            col_idx = columns.index('labelIds') + 1
            color = v['labelColors'][0].replace('#', '') if v['labelColors'] else None
            if color:
                cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore
                cell.font = Font(color=color)  # type: ignore
        # Row coloring based on hu
        if colored and 'hu' in columns:
            hu = v.get('hu')
            color = get_hu_color(hu) if hu else None
            if color:
                for col_idx in range(1, len(columns)+1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")  # type: ignore

    # Auto-size columns
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):  # type: ignore
        max_length = 0
        col_idx_val = getattr(col[0], 'column', 1) if col else 1
        if not isinstance(col_idx_val, int):
            col_idx_val = 1
        col_letter = get_column_letter(col_idx_val)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # type: ignore

    filename = f'{base_name}_{datetime.now().date().isoformat()}.xlsx'
    wb.save(filename)
    print(f"Excel file saved as {filename}")

def main():
    args = parse_args()
    csv_file = args.file
    if not os.path.exists(csv_file):
        print(f"CSV file '{csv_file}' not found.")
        sys.exit(1)
    vehicles = send_csv_to_server(csv_file)
    vehicles = sort_vehicles(vehicles)
    write_excel(vehicles, args.keys, args.colored, csv_file)

if __name__ == "__main__":
    main()
